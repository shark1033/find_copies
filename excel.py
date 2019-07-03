import time

import openpyxl
from openpyxl.styles import PatternFill


class ExcelOpen:


    def open_file(self, file_name,list_name,column):
        self.col=column
        self.file_name=file_name
        elist=list_name
        self.wb = openpyxl.load_workbook(filename=self.file_name)
        self.sheet = self.wb.get_sheet_by_name(list_name)
        print(self.sheet)

    def open_file2(self,):

        self.wb = openpyxl.load_workbook(filename='C:\\Users\\shark10\\Desktop\\сортировка.xlsx')
        self.sheet = self.wb.get_sheet_by_name("наше")
        print(self.sheet)

    def read_and_write(self):

        print("read and write")
        counter=0
        for i in range(1, self.sheet.max_row):
            for y in range(i+1, self.sheet.max_row):
                if self.sheet.cell(row=i, column=self.col).value == self.sheet.cell(row=y, column=1).value:
                    #self.sheet.cell(row=y, column=2).value = 'Повтор!'
                    #self.sheet.cell(row=i, column=2).value = 'Повтор!'
                    self.sheet.cell(row=i, column=self.col).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
                    self.sheet.cell(row=y, column=self.col).fill = PatternFill(start_color='FFEE08', end_color='FFEE08',
                                                                        fill_type='solid')


    def save_file(self):
        print("save")
        newFile=self.file_name
        newarray=newFile.split('.')
        print(newarray[0])
        mainpart=newarray[0]
        now=time.strftime("%B_%d_%Y", time.localtime())
        print(now)
        print(type(now))
        output=mainpart + "_" + now+".xlsx"
        print(output)
        self.wb.save(output)

    def save_file2(self):
        print("save")

        now=time.strftime("%B_%d_%Y", time.localtime())
        output="C:\\Users\\shark10\\Desktop\\сортировка" + "_" + now + ".xlsx"
        print(output)
        self.wb.save(output)

# Another program


    def count(self):
        print("func count")

        self.sheet.cell(row=1, column=4).value=self.sheet.cell(row=1, column=1).value
        self.sheet.cell(row=1, column=5).value = self.sheet.cell(row=1, column=2).value

        for i in range(1, 331):  #check
            print(i)
            counter=0
            counter_str=0
            print(self.sheet.cell(row=i, column=1).value)
            if self.check_rep(self.sheet.cell(row=i, column=1).value):
                print("True")
                continue
            else:
                print("Else")
                for y in range(i+1, 331):
                    if self.sheet.cell(row=i, column=1).value==self.sheet.cell(row=y, column=1).value:
                       counter+=self.sheet.cell(row=y, column=2).value
                       counter_str+=1
                       print("counter:" +str(counter))
                       print(counter_str)

            self.sheet.cell(row=i, column=4).value = self.sheet.cell(row=i, column=1).value
            if counter_str==0:
                self.sheet.cell(row=i, column=5).value = self.sheet.cell(row=i, column=2).value
            else:
                self.sheet.cell(row=i, column=5).value = counter + self.sheet.cell(row=i, column=2).value

    def check_rep(self, item) :  # проверяем нет ли уже записи в новой колонке
        print("func check")
        for z in range(1, 331):
            print(item)
            print(self.sheet.cell(row=z, column=4).value)
            if item == self.sheet.cell(row=z, column=4).value:
                return True
        return False
